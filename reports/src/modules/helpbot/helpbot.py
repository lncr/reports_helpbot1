import os
import telebot
import constants
import constants_ru
import openpyxl
from telebot import types

from .utils import create_lang_inlines
from fastapi import APIRouter, Request

TOKEN = constants.TOKEN
bot = telebot.TeleBot(TOKEN, parse_mode='HTML')
user_dict = {}
lang_dict = {}
history_dict = {}

bot.remove_webhook()
bot.set_webhook(url='https://it-licey-chat-bot-9cbba0442416.herokuapp.com/api/v1/'+TOKEN)

router = APIRouter()


def write_to_xlsx(user):
	wb = openpyxl.load_workbook(filename='stats.xlsx')
	ws = wb.worksheets[0]
	max_row = ws.max_row
	for i in range(2, ws.max_row+1):
		if ws.cell(row=i, column=1).value == user.telegram_id:
			ws[f'A{i}'] = user.telegram_id
			ws[f'B{i}'] = user.telegram_username
			ws[f'C{i}'] = user.name
			ws[f'D{i}'] = user.credential
			break
	else:
		ws[f'A{max_row+1}'] = user.telegram_id
		ws[f'B{max_row+1}'] = user.telegram_username
		ws[f'C{max_row+1}'] = user.name
		ws[f'D{max_row+1}'] = user.credential
	wb.save(filename='stats.xlsx')


buttons = {
	'ru': {
		'main_menu': 'Главное меню',
		'glossary': 'Показать все вопросы',
		'langs': 'Языки программирования',
		'ages': 'Ограничения по возрасту',
		'wanna_langs': 'Языки программирования',
		'ort': 'Нужен ли ОРТ?',
		'contacts': 'Контакты лицеев',
		'education_program': 'Программа обучения',
		'education_duration': 'Длительность обучения',
		'prev_menu': 'Назад',
		'contract_group': 'Контрактные группы',
		'budget_group': 'Бюджетные группы',
		'documents': 'Документы',
		'laptop': 'Характеристики ноутбука',
		'dorm': 'Общежитие'
	},
	'kg': {
		'main_menu': 'Башкы  меню',
		'glossary': 'Бардык суроолорду көрсөтүү',
		'langs': 'Кандай программалоо тилдерин үйрөнсөм болот?',
		'ages': 'Жашы боюнча чектөөлөр барбы?',
		'wanna_langs': 'Кандай программалоо тилдерин үйрөнсөм болот?',
		'ort': 'Тапшыруу үчүн ЖРТ керекпи?',
		'contacts': 'Байланыш номерлер',
		'education_program': 'Окутуунун программасы',
		'education_duration': 'Окутуунун мөөнөтү',
		'prev_menu': 'Артка кайтуу',
		'contract_group': 'Келишим топтору',
		'budget_group': 'Бюджеттик топтор',
		'documents': 'Керектүү документтер',
		'laptop': 'Кандай ноутбук талап кылынат?',
		'dorm': 'Лицейлерде жатакана барбы?'
	}
}

def create_content(language=None):
	language = 'kg' if not language else language
	
	main_menu = types.InlineKeyboardButton(text=buttons[language]['main_menu'], callback_data='main_menu')
	glossary = types.InlineKeyboardButton(text=buttons[language]['glossary'], callback_data='glossary')
	langs = types.InlineKeyboardButton(text=buttons[language]['langs'], callback_data='languages')
	ages = types.InlineKeyboardButton(text=buttons[language]['ages'], callback_data='ages')
	ort = types.InlineKeyboardButton(text=buttons[language]['ort'], callback_data='ort')
	wanna_langs = types.InlineKeyboardButton(text=buttons[language]['wanna_langs'], callback_data='languages')
	contacts = types.InlineKeyboardButton(text=buttons[language]['contacts'], callback_data='contacts')
	python = types.InlineKeyboardButton(text='Python', callback_data='python')
	javascript = types.InlineKeyboardButton(text='JavaScript', callback_data='javascript')
	education_program = types.InlineKeyboardButton(text=buttons[language]['education_program'], callback_data='education_program')
	education_duration = types.InlineKeyboardButton(text=buttons[language]['education_duration'], callback_data='education_duration')
	prev_menu = types.InlineKeyboardButton(text=buttons[language]['prev_menu'], callback_data='previous')
	contract_group = types.InlineKeyboardButton(text=buttons[language]['contract_group'], callback_data='contract_group')
	budget_group = types.InlineKeyboardButton(text=buttons[language]['budget_group'], callback_data='budget_group')
	documents = types.InlineKeyboardButton(text=buttons[language]['documents'], callback_data='documents')
	laptop = types.InlineKeyboardButton(text=buttons[language]['laptop'], callback_data='laptop')
	dorm = types.InlineKeyboardButton(text=buttons[language]['dorm'], callback_data='dorm')

	main_txt = 'Баштайлы! Бир суроону тандаңыз' if language == 'kg' else 'Начнем! Что тебя интересует:'
	glossary_txt = 'Эмне жөнүндө маалымат издеп жатасыз?' if language == 'kg' else 'О чём вы хотите узнать?'

	if language == 'ru':
		used_constants = constants_ru
	else:
		used_constants = constants

	CONTENT = {
		'main_menu': (main_txt, [[langs], [ages], [ort], [contacts], [glossary]]),
		'glossary': (glossary_txt, [[python], [javascript], [education_program], [education_duration], 
						  			[contract_group], [budget_group], [documents], [laptop], [dorm], ]),
		'contacts': (used_constants.CONTACTS, [[main_menu]]),
		'languages': (used_constants.LANGUAGES, [[python], [javascript], [main_menu]]),
		'ages': (used_constants.AGES, [[main_menu]]),
		'ort': (used_constants.ORT, [[main_menu]]),
		'javascript': (used_constants.AGE_BASED, create_lang_inlines('javascript', language)),
		'python': (used_constants.AGE_BASED, create_lang_inlines('python', language)),
		'class_9_javascript': (used_constants.CLASS_9_JAVASCRIPT, [[education_program], [education_duration], [main_menu]]),
		'class_11_javascript': (used_constants.CLASS_11_JAVASCRIPT, [[education_program], [education_duration], [main_menu]]),
		'class_9_python': (used_constants.CLASS_9_PYTHON, [[education_program], [education_duration], [main_menu]]),
		'class_11_python': (used_constants.CLASS_11_PYTHON, [[education_program], [education_duration], [main_menu]]),
		'education_program': (used_constants.EDUCATION_PROGRAM, [[main_menu]]),
		'education_duration': (used_constants.EDUCATION_DURATION, [[contract_group], [budget_group], [main_menu]]),
		'contract_group': (used_constants.CONTRACT_GROUP, [[documents], [laptop], [dorm], [main_menu]]),
		'budget_group': (used_constants.BUDGET_GROUP, [[documents], [laptop], [dorm], [main_menu]]),
		'documents': (used_constants.DOCUMENTS, [[main_menu]]),
		'laptop': (used_constants.LAPTOP, [[main_menu]]),
		'dorm': (used_constants.DORM, [[main_menu]])
	}
	return CONTENT


class User:
	def __init__(self, name, telegram_username=None, telegram_id=None):
		self.name = name
		self.telegram_username = telegram_username
		self.telegram_id = telegram_id
		self.credential = None

def process_name_step(message):
	chat_id = message.chat.id
	name = message.text
	user = User(name, message.from_user.username, message.from_user.id)
	user_dict[chat_id] = user

	if lang_dict[chat_id] == 'kg':
		reply = 'Номериңизди же почтаңызды калтырсаңыз болот, лицейлерге кабыл алуу башталганда, билдирүүнү жөнөтөбүз'
	else:
		reply = 'Оставь свой номер или почту, отправим уведомление о старте набора в группы'
	msg = bot.reply_to(message, reply)
	bot.register_next_step_handler(msg, process_credential_step)


def process_credential_step(message):
	chat_id = message.chat.id
	credential = message.text
	user = user_dict[chat_id]
	user.credential = credential
	write_to_xlsx(user)
	CONTENT = create_content(lang_dict[chat_id])
	reply_markup = types.InlineKeyboardMarkup(CONTENT['main_menu'][1])
	bot.send_message(message.chat.id, CONTENT['main_menu'][0], reply_markup=reply_markup)


@bot.message_handler(commands=['start'])
def send_welcome(message):
	kg = types.InlineKeyboardButton(text='Кыргызча', callback_data='set_kg_lng')
	ru = types.InlineKeyboardButton(text='Русский', callback_data='set_ru_lng')
	reply_markup = types.InlineKeyboardMarkup([[kg], [ru]])

	history = history_dict.get(message.chat.id)
	if history:
		history_dict[message.chat.id] = []

	bot.send_message(message.chat.id, constants.INITIAL_MSG, reply_markup=reply_markup)


@bot.message_handler(commands=['loadstats'])
def load_stats(message):
	doc = open('stats.xlsx', 'rb')
	bot.send_document(message.chat.id, doc)


@bot.callback_query_handler(func=lambda call:True)
def answer(call):

	history = history_dict.get(call.message.chat.id, [])

	CONTENT = create_content(lang_dict.get(call.message.chat.id))

	if call.data == 'previous':
		history.pop()
		content = CONTENT[history[-1]] if history else CONTENT['main_menu']
		keyboards = content[1]

		if history and not len(keyboards[-1]) > 1:
			txt = buttons.get(lang_dict.get(call.message.chat.id))['prev_menu']
			prev_menu = types.InlineKeyboardButton(text=txt, callback_data='previous')
			keyboards[-1].insert(0, prev_menu)

		reply_markup = types.InlineKeyboardMarkup(keyboards)
		bot.send_message(call.message.chat.id, content[0], reply_markup=reply_markup)

	elif call.data == 'main_menu' or call.data == 'glossary':
		history = []
		reply_markup = types.InlineKeyboardMarkup(CONTENT[call.data][1])
		
		bot.send_message(call.message.chat.id, CONTENT[call.data][0], reply_markup=reply_markup)
	
	elif call.data == 'set_kg_lng' or call.data == 'set_ru_lng':
		chat_id = call.message.chat.id
		lang_dict[chat_id] = call.data[4:6]
		lang = lang_dict.get(call.message.chat.id)
		if lang == 'ru':
			from constants_ru import WELCOME_MSG
			reply = WELCOME_MSG
		else:
			reply = constants.WELCOME_MSG
		msg = bot.reply_to(call.message, reply)
		bot.register_next_step_handler(msg, process_name_step)

	else:
		content = CONTENT[call.data]
		text = content[0]
		keyboards = content[1]

		if history and not len(keyboards[-1]) > 1:
			txt = buttons.get(lang_dict.get(call.message.chat.id))['prev_menu']
			prev_menu = types.InlineKeyboardButton(text=txt, callback_data='previous')
			keyboards[-1].insert(0, prev_menu)

		reply_markup = types.InlineKeyboardMarkup(keyboards)
		history.append(call.data)
		bot.send_message(call.message.chat.id, text, reply_markup=reply_markup)

	history_dict[call.message.chat.id] = history
	bot.answer_callback_query(call.id)



@router.post(f"/{TOKEN}")
async def get_message(request: Request):
    json_data = await request.json()
    update = types.Update.de_json(json_data)
    bot.process_new_updates([update])
    return {"status": "ok"}
